from rest_framework import generics, permissions, mixins, status
from django_filters.rest_framework import DjangoFilterBackend
from rest_framework.parsers import MultiPartParser
from .models import Project, ImportacaoProjeto
from rest_framework.response import Response
from .serializers import ProjectSerializer
from rest_framework.views import APIView
from django.shortcuts import get_object_or_404
from django.shortcuts import get_object_or_404, render, redirect
from .filters import ProjectFilter
from .services import importar_planilha_projetos, registrar_log_status
from django.utils import timezone

class ProjectCreateAPIView(generics.CreateAPIView):
    queryset = Project.objects.all()
    serializer_class = ProjectSerializer
    permission_classes = [permissions.IsAdminUser]

    def perform_create(self, serializer):
        projeto = serializer.save(criado_por=self.request.user.nome + " (" + self.request.user.email + ")")
       
        registrar_log_status(
            projeto=projeto,
            status_anterior=None,  
            status_novo=projeto.status,
            usuario=self.request.user
        )

class ProjectRetrieveAPIView(generics.RetrieveAPIView):
    queryset = Project.objects.all()
    serializer_class = ProjectSerializer
    permission_classes = [permissions.IsAuthenticated]
    lookup_field = 'id'  

    def get_queryset(self):
        user = self.request.user
        queryset = super().get_queryset()

        if not user.is_staff:
            queryset = queryset.filter(ativo=True)

        return queryset

class ProjectListAPIView(generics.ListAPIView):
    serializer_class = ProjectSerializer
    permission_classes = [permissions.IsAuthenticated]
    filter_backends = [DjangoFilterBackend]
    filterset_class = ProjectFilter

    def get_queryset(self):
        user = self.request.user
        queryset = Project.objects.all()
        
        if not user.is_staff:
            queryset = queryset.filter(ativo=True)

        return queryset

class ProjectUpdateAPIView(generics.UpdateAPIView):
    queryset = Project.objects.all()
    serializer_class = ProjectSerializer
    permission_classes = [permissions.IsAdminUser]
    
    def perform_update(self, serializer):
        projeto_antigo = self.get_object()
        status_antigo = projeto_antigo.status

        projeto = serializer.save(atualizado_por=f"{self.request.user.nome} ({self.request.user.email})")
        status_novo = projeto.status

        registrar_log_status(projeto, status_antigo, status_novo, self.request.user)

class ProjectDeleteAPIView(generics.DestroyAPIView):
    queryset = Project.objects.all()
    serializer_class = ProjectSerializer
    permission_classes = [permissions.IsAdminUser]

class ProjectBulkDeleteAPIView(mixins.DestroyModelMixin, generics.GenericAPIView):
    queryset = Project.objects.all()
    permission_classes = [permissions.IsAdminUser]

    def delete(self, request, *args, **kwargs):
        ids = request.data.get('ids', [])
        if not isinstance(ids, list):
            return Response({"error": "ids precisa ser uma lista"}, status=status.HTTP_400_BAD_REQUEST)

        projects = self.get_queryset().filter(id__in=ids)
        count = projects.count()
        projects.delete()
        return Response({"message": f"{count} projetos deletados."}, status=status.HTTP_200_OK)
    
class ImportarProjetosView(APIView):
    parser_classes = [MultiPartParser]
    permission_classes = [permissions.IsAdminUser]

    def post(self, request, *args, **kwargs):
        arquivo = request.FILES.get('arquivo')

        print("Arquivo recebido:", arquivo)
    
        if not arquivo:
            return Response({"erro": "Arquivo não enviado."}, status=status.HTTP_400_BAD_REQUEST)

        importacao = ImportacaoProjeto.objects.create(arquivo=arquivo)
        try:
            importar_planilha_projetos(importacao, self.request)
        except Exception as e:
            return Response({"erro": str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

        return Response({
            "mensagem": "Importação realizada com sucesso.",
            "linhas_lidas": importacao.linhas_lidas,
            "projetos_criados": importacao.projetos_criados,
            "projetos_ignorados": importacao.projetos_ignorados,
            "linhas_ignoradas": importacao.linhas_ignoradas_texto,
            "id": importacao.id
        })
    
class VerificarInscricaoView(APIView):
    permission_classes = [permissions.IsAuthenticated]
    def get(self, request, project_id):
        projeto = get_object_or_404(Project, id=project_id)
        agora = timezone.now()

        pode_inscrever = projeto.inicio_inscricoes <= agora <= projeto.fim_inscricoes
        return Response({"pode_inscrever": pode_inscrever})

from applications.models import Application

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from django.http import HttpResponse
from django.utils.text import slugify

def exportar_alunas_excel(request, projeto_id):
    try:
        if isinstance(projeto_id, str):
            projeto_id = uuid.UUID(projeto_id)
    except (ValueError, AttributeError):
        pass
    
    projeto = get_object_or_404(Project, id=projeto_id)
    
    # Puxa todas as aplicações com os relacionamentos necessários
    alunas = (
        Application.objects
        .filter(projeto=projeto)
        .select_related(
            'usuario__escola__tipo_ensino',
            'tipo_de_vaga'
        )
    )

    # Cria uma workbook e worksheet
    wb = openpyxl.Workbook()
    ws = wb.active
    
    # Título seguro para a planilha
    ws.title = "Alunas Vinculadas"

    # Cabeçalhos
    headers = [
        'Nome', 
        'Tipo de Vaga', 
        'Tipo de Ensino', 
        'Status', 
        'Nota Final (%)'
    ]
    
    # Estilo do cabeçalho
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")

    # Adiciona cabeçalhos
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment

    # Adiciona dados
    row_num = 2
    for a in alunas:
        historico = HistoricoEscolar.objects.filter(usuario=a.usuario).first()

        # Calcula a nota final percentual
        nota_final_percentual = "-"
        if historico:
            notas = historico.notas.all()
            if notas.exists():
                total = sum([n.nota_final_percentual or 0 for n in notas])
                media = total / len(notas)
                nota_final_percentual = round(media, 2)

        # Dados da linha
        row_data = [
            a.usuario.nome or a.usuario.username,
            getattr(a.tipo_de_vaga, "nome", "-"),
            getattr(a.usuario.escola.tipo_ensino, "nome", "-"),
            a.get_status_display(),
            nota_final_percentual,
        ]

        # Adiciona linha
        for col_num, value in enumerate(row_data, 1):
            ws.cell(row=row_num, column=col_num, value=value)

        row_num += 1

    # Ajusta largura das colunas
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width

    # Adiciona informações do projeto como primeira linha
    ws.insert_rows(1)
    ws.merge_cells('A1:E1')
    project_info_cell = ws.cell(row=1, column=1, value=f"Projeto: {projeto.nome} - Exportado em {timezone.now().strftime('%d/%m/%Y %H:%M')}")
    project_info_cell.font = Font(bold=True, size=12)
    project_info_cell.alignment = Alignment(horizontal="center")

    # Cria resposta HTTP
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    filename = f"alunas_{slugify(projeto.nome)}_{timezone.now().strftime('%Y-%m-%d')}.xlsx"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    
    wb.save(response)
    return response

def detalhes_projeto(request, projeto_id):
    projeto = get_object_or_404(Project, id=projeto_id)

    # Puxa todas as aplicações com os relacionamentos necessários
    alunas = (
        Application.objects
        .filter(projeto=projeto)
        .select_related(
            'usuario__escola__tipo_ensino',
            'tipo_de_vaga'
        )
    )

    # Cria uma lista para enviar ao template
    alunas_data = []
    for a in alunas:
        historico = HistoricoEscolar.objects.filter(usuario=a.usuario).first()

        # Calcula a nota final percentual (média das notas normalizadas)
        nota_final_percentual = "-"
        if historico:
            notas = historico.notas.all()
            if notas.exists():
                total = sum([n.nota_final_percentual or 0 for n in notas])
                media = total / len(notas)
                nota_final_percentual = round(media, 2)

        alunas_data.append({
            "id": a.id,
            "nome": a.usuario.nome or a.usuario.username,
            "tipo_de_vaga": getattr(a.tipo_de_vaga, "nome", "-"),
            "tipo_ensino": getattr(a.usuario.escola.tipo_ensino, "nome", "-"),
            "status_label": a.get_status_display(),
            "status_value": a.status,
            "nota_final_percentual": nota_final_percentual,
        })

    context = {
        'projeto': projeto,
        'alunas': alunas_data,
        'status_choices': Application.STATUS_ESCOLHAS,
    }

    return render(request, 'components/projects/detalhes_projeto.html', context)


from django.views.decorators.csrf import csrf_exempt
from django.http import JsonResponse
import json

@csrf_exempt
def atualizar_status(request, pk):
    if request.method == 'POST':
        data = json.loads(request.body)
        novo_status = data.get('status')

        if novo_status not in dict(Application.STATUS_ESCOLHAS):
            return JsonResponse({'error': 'Status inválido'}, status=400)

        try:
            app = Application.objects.get(pk=pk)
            app.status = novo_status
            app.save()
            return JsonResponse({'success': True, 'status': novo_status})
        except Application.DoesNotExist:
            return JsonResponse({'error': 'Registro não encontrado'}, status=404)
    return JsonResponse({'error': 'Método não permitido'}, status=405)

from users.models.school_transcript_model import HistoricoEscolar


def perfil_aluna(request, pk):
    aplicacao = get_object_or_404(Application, pk=pk)
    usuario = aplicacao.usuario

    # tenta pegar o histórico da aluna, se existir
    historico = (
        HistoricoEscolar.objects.filter(usuario=usuario).first()
    )

    # pega as notas ordenadas por disciplina e bimestre
    notas = []
    if historico:
        notas = (
            historico.notas
            .select_related('disciplina')
            .order_by('disciplina__nome', 'bimestre')
        )

    contexto = {
        'aplicacao': aplicacao,
        'usuario': usuario,
        'historico': historico,
        'notas': notas,
    }
    return render(request, 'components/projects/perfil_aluna.html', contexto)



from django.shortcuts import render
from django.shortcuts import render
from .models import Project, Regiao, Estado
from django.db.models import Count
from django.core.paginator import Paginator

def lista_projetos(request):
    queryset = Project.objects.all()

    q = request.GET.get('q', '')
    regiao_id = request.GET.get('regiao', '')
    estado_id = request.GET.get('estado', '')  
    formato = request.GET.get('formato', '')
    ordenar = request.GET.get('ordenar', 'nome')

    if q:
        queryset = queryset.filter(nome__icontains=q)

    if regiao_id:
        queryset = queryset.filter(regioes_aceitas__id=regiao_id)

    if estado_id:
        queryset = queryset.filter(estados_aceitos__id=estado_id) 

    if formato:
        queryset = queryset.filter(formato=formato)

    allowed_order_fields = ['nome', 'data_inicio', 'data_fim', 'vagas']
    if ordenar not in allowed_order_fields:
        ordenar = 'nome'

    queryset = queryset.order_by(ordenar)

    # Paginação
    from django.core.paginator import Paginator
    paginator = Paginator(queryset, 10)
    page_number = request.GET.get('page')
    projetos_page = paginator.get_page(page_number)

    regioes = Regiao.objects.all()
    estados = Estado.objects.all()  

    return render(request, 'components/projects/lista_projetos.html', {
        'projetos': projetos_page,
        'regioes': regioes,
        'estados': estados,  
        'filtros': {
            'q': q,
            'regiao': regiao_id,
            'estado': estado_id, 
            'formato': formato,
            'ordenar': ordenar,
        }
    })

def lista_projetos_para_homologar_inscricao(request):

    queryset = Project.objects.annotate(
        num_inscritos=Count('application') 
    )
    
    q = request.GET.get('q', '')
    regiao_id = request.GET.get('regiao', '')
    estado_id = request.GET.get('estado', '')  
    formato = request.GET.get('formato', '')
    ordenar = request.GET.get('ordenar', 'nome')

    if q:
        queryset = queryset.filter(nome__icontains=q)

    if regiao_id:
        queryset = queryset.filter(regioes_aceitas__id=regiao_id)

    if estado_id:
        queryset = queryset.filter(estados_aceitos__id=estado_id) 

    if formato:
        queryset = queryset.filter(formato=formato)

    allowed_order_fields = ['nome', 'data_inicio', 'data_fim', 'vagas', 'num_inscritos', '-num_inscritos']
    if ordenar not in allowed_order_fields and ordenar.strip('-') not in allowed_order_fields:
        ordenar = 'nome'

    queryset = queryset.order_by(ordenar)

    # Paginação
    paginator = Paginator(queryset, 10)
    page_number = request.GET.get('page')
    projetos_page = paginator.get_page(page_number)

    regioes = Regiao.objects.all()
    estados = Estado.objects.all()  

    return render(request, 'components/projects/lista_projetos_para_avaliacao.html', {
        'projetos': projetos_page,
        'regioes': regioes,
        'estados': estados,  
        'filtros': {
            'q': q,
            'regiao': regiao_id,
            'estado': estado_id, 
            'formato': formato,
            'ordenar': ordenar,
        }
    })


